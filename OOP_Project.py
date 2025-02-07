import openpyxl
from datetime import datetime
import os

# --- Classes ---

class Wallet:
    def __init__(self, balance=0.0):
        self.balance = balance

    def deposit(self, amount):
        self.balance += amount

    def withdraw(self, amount):
        if self.balance >= amount:
            self.balance -= amount
            return True
        return False

    def check_balance(self):
        return self.balance


class User:
    def __init__(self, user_id, name, phone_number, wallet=None):
        self.user_id = user_id
        self.name = name
        self.phone_number = phone_number
        self.wallet = wallet if wallet else Wallet()

    def display_details(self):
        return f"User ID: {self.user_id}, Name: {self.name}, Phone: {self.phone_number}"

    def receive_money(self, amount):
        self.wallet.deposit(amount)


class Transaction:
    def __init__(self, transaction_id, sender, receiver, amount, date=None):
        self.transaction_id = transaction_id
        self.sender = sender
        self.receiver = receiver
        self.amount = amount
        self.date = date if date else datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def to_dict(self):
        dict = {
            "Transaction ID": self.transaction_id,
            "Sender": self.sender.user_id,
            "Receiver": self.receiver.user_id,
            "Amount": self.amount,
            "Date": self.date,
        }
        return dict


# --- Helper Functions ---

def load_users(filename="users.xlsx"):
    users = {}
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["User ID", "Name", "Phone Number", "Balance"])
        wb.save(filename)
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        user_id, name, phone, balance = row
        users[user_id] = User(user_id, name, phone, Wallet(balance))
    return users


def save_users(users, filename="users.xlsx"):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["User ID", "Name", "Phone Number", "Balance"])
    for user in users.values():
        sheet.append([user.user_id, user.name, user.phone_number, user.wallet.check_balance()])
    wb.save(filename)



def load_transactions(filename="transactions.xlsx"):
    transactions = []
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Transaction ID", "Sender ID", "Receiver ID", "Amount", "Date"])
        wb.save(filename)
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        transactions.append(row)
    return transactions


def save_transactions(transactions, filename="transactions.xlsx"):
    # Check if the file exists, if not, create it with headers
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Transaction ID", "Sender ID", "Receiver ID", "Amount", "Date"])
        wb.save(filename)

    # Load the existing file
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Get the existing transaction IDs to avoid duplicates
    existing_ids = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True)}

    # Get the last transaction ID and increment it for new transactions
    last_transaction_id = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        transaction_id = row[0]
        # Assuming Transaction ID is prefixed with 'T' and followed by a number
        if transaction_id.startswith('T'):
            try:
                last_transaction_id = max(last_transaction_id, int(transaction_id[1:]))
            except ValueError:
                pass

    # Append new transactions with auto-incremented transaction IDs
    for t in transactions:
        # Increment the transaction ID
        last_transaction_id += 1
        new_transaction_id = f"T{last_transaction_id:03d}"

        if new_transaction_id not in existing_ids:
            sheet.append([new_transaction_id, t.sender.user_id, t.receiver.user_id, t.amount, t.date])
            existing_ids.add(new_transaction_id)  # Add the new transaction ID to existing_ids set

    # Save the updated file
    wb.save(filename)





# --- Main Program ---

def main():
    users = load_users()
    transactions = []
    
    print("\nWelcome to the Mobile Payment System!")


    while True:
    
        print("1. Register New User")
        print("2. Check Balance")
        print("3. Send Money")
        print("4. Receive Money")
        print("5. View Transaction History")
        print("6. Exit")
        choice = input("Choose an option: ")

        if choice == "1":
            user_id = input("Enter User ID: ")
            name = input("Enter Name: ")
            phone = input("Enter Phone Number: ")
            balance = float(input("Enter Initial Balance: "))
            if user_id in users:
                print("User ID already exists.")
            else:
                users[user_id] = User(user_id, name, phone, Wallet(balance))
                save_users(users)
                print("User registered successfully!")

        elif choice == "2":
            user_id = input("Enter User ID: ")
            if user_id in users:
                balance = users[user_id].wallet.check_balance()
                print(f"Current Balance: ${balance:.2f}")
            else:
                print("User not found.")

        elif choice == "3":
            sender_id = input("Enter Sender User ID: ")
            receiver_id = input("Enter Receiver User ID: ")
            amount = float(input("Enter Amount: "))
            if sender_id in users and receiver_id in users:
                sender = users[sender_id]
                receiver = users[receiver_id]
                if sender.wallet.withdraw(amount):
                    receiver.receive_money(amount)
                    transaction_id = f"T{len(transactions)+1:03d}"
                    transaction = Transaction(transaction_id, sender, receiver, amount)
                    transactions.append(transaction)
                    save_users(users)  # Save updated balances
                    #save_transactions(transactions)  # Save the transaction
                    print(f"Transaction successful! {amount:.2f} sent to {receiver.name}.")
                else:
                    print("Insufficient balance. Transaction failed.")
            else:
                print("Sender or Receiver not found.")

        elif choice == "4":
            user_id = input("Enter User ID : ")
            receiver_id = input("Enter Receiver User ID : ")
            amount = float(input("Enter Amount: "))
            if user_id in users and receiver_id in users:
                users[user_id].receive_money(amount)
                save_users(users)  # Save updated balance
                print(f"Transaction successful! ${amount:.2f} received.")
            else:
                print("User not found.")

        elif choice == "5":
            print("Transaction History:")
            for t in transactions:
                print(f"{t.sender.user_id} sent ${t.amount:.2f} to {t.receiver.user_id} on {t.date}")

        elif choice == "6":
            save_users(users)
            save_transactions(transactions)
            print("Exiting... Goodbye!")
            break

        else:
            print("Invalid option. Please try again.")


if __name__ == "__main__":
    main()
