from abc import ABC, abstractmethod
import openpyxl
from datetime import datetime
import os


class WalletInterface(ABC):
    """Abstract class defining the wallet interface."""

    @abstractmethod
    def deposit(self, amount):
        pass

    @abstractmethod
    def withdraw(self, amount):
        pass

    @abstractmethod
    def check_balance(self):
        pass


class Wallet(WalletInterface):

    def __init__(self, balance=0.0):
        self.balance = balance

    def deposit(self, amount):
        self.balance += amount

    def withdraw(self, amount):
        if self.balance >= amount:
            self.balance -= amount
            return True
        return False

    def check_balance(self):
        return self.balance


class UserInterface(ABC):
    """Abstract class defining the user interface."""

    @abstractmethod
    def display_details(self):
        pass

    @abstractmethod
    def receive_money(self, amount):
        pass


class User(UserInterface):
    """Concrete implementation of the UserInterface."""

    def __init__(self, user_id, name, phone_number, wallet=None):
        self.user_id = user_id
        self.name = name
        self.phone_number = phone_number
        self.wallet = wallet if wallet else Wallet()

    def display_details(self):
        return f"User ID: {self.user_id}, Name: {self.name}, Phone: {self.phone_number}"

    def receive_money(self, amount):
        self.wallet.deposit(amount)


class Transaction:
    """Represents a transaction between two users."""

    def __init__(self, transaction_id, sender, receiver, amount, date=None):
        self.transaction_id = transaction_id
        self.sender = sender
        self.receiver = receiver
        self.amount = amount
        self.date = date if date else datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def to_dict(self):
        return {
            "Transaction ID": self.transaction_id,
            "Sender": self.sender.user_id,
            "Receiver": self.receiver.user_id,
            "Amount": self.amount,
            "Date": self.date,
        }


class PaymentSystemInterface(ABC):
    """Abstract class defining the payment system interface."""

    @abstractmethod
    def register_user(self):
        pass

    @abstractmethod
    def check_balance(self):
        pass

    @abstractmethod
    def send_money(self):
        pass

    @abstractmethod
    def view_transactions(self):
        pass


class MobilePaymentSystem(PaymentSystemInterface):
    """Concrete implementation of the PaymentSystemInterface."""

    def __init__(self, users_file="users.xlsx", transactions_file="transactions.xlsx"):
        self.users_file = users_file
        self.transactions_file = transactions_file
        self.users = self.load_users()
        self.transactions = self.load_transactions()

    def load_users(self):
        users = {}
        if not os.path.exists(self.users_file):
            self._initialize_file(self.users_file, ["User ID", "Name", "Phone Number", "Balance"])
        wb = openpyxl.load_workbook(self.users_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            user_id, name, phone, balance = row
            users[user_id] = User(user_id, name, phone, Wallet(balance))
        return users

    def save_users(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["User ID", "Name", "Phone Number", "Balance"])
        for user in self.users.values():
            sheet.append([user.user_id, user.name, user.phone_number, user.wallet.check_balance()])
        wb.save(self.users_file)

    def load_transactions(self):
        transactions = []
        if not os.path.exists(self.transactions_file):
            self._initialize_file(self.transactions_file, ["Transaction ID", "Sender ID", "Receiver ID", "Amount", "Date"])
        wb = openpyxl.load_workbook(self.transactions_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            transaction_id, sender_id, receiver_id, amount, date = row
            sender = self.users.get(sender_id)
            receiver = self.users.get(receiver_id)
            if sender and receiver:
                transactions.append(Transaction(transaction_id, sender, receiver, amount, date))
        return transactions

    def save_transactions(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Transaction ID", "Sender ID", "Receiver ID", "Amount", "Date"])
        for t in self.transactions:
            sheet.append([t.transaction_id, t.sender.user_id, t.receiver.user_id, t.amount, t.date])
        wb.save(self.transactions_file)

    @staticmethod
    def _initialize_file(filename, headers):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(headers)
        wb.save(filename)

    def register_user(self):
        user_id = input("Enter User ID: ")
        name = input("Enter Name: ")
        phone = input("Enter Phone Number: ")
        balance = float(input("Enter Initial Balance: "))
        if user_id in self.users:
            print("User ID already exists.")
        else:
            self.users[user_id] = User(user_id, name, phone, Wallet(balance))
            self.save_users()
            print("User registered successfully!")

    def check_balance(self):
        user_id = input("Enter User ID: ")
        if user_id in self.users:
            balance = self.users[user_id].wallet.check_balance()
            print(f"Current Balance: ${balance:.2f}")
        else:
            print("User not found.")

    def send_money(self):
        sender_id = input("Enter Sender User ID: ")
        receiver_id = input("Enter Receiver User ID: ")
        amount = float(input("Enter Amount: "))
        if sender_id in self.users and receiver_id in self.users:
            sender = self.users[sender_id]
            receiver = self.users[receiver_id]
            if sender.wallet.withdraw(amount):
                receiver.receive_money(amount)
                transaction_id = f"T{len(self.transactions) + 1:03d}"
                transaction = Transaction(transaction_id, sender, receiver, amount)
                self.transactions.append(transaction)
                self.save_users()
                self.save_transactions()
                print(f"Transaction successful! ${amount:.2f} sent to {receiver.name}.")
            else:
                print("Insufficient balance. Transaction failed.")
        else:
            print("Sender or Receiver not found.")

    def view_transactions(self):
        print("Transaction History:")
        for t in self.transactions:
            print(f"{t.sender.name} sent ${t.amount:.2f} to {t.receiver.name} on {t.date}")

    def run(self):
        while True:
            print("\nWelcome to the Mobile Payment System!")
            print("1. Register New User")
            print("2. Check Balance")
            print("3. Send Money")
            print("4. View Transactions")
            print("5. Exit")
            choice = input("Choose an option: ")

            if choice == "1":
                self.register_user()
            elif choice == "2":
                self.check_balance()
            elif choice == "3":
                self.send_money()
            elif choice == "4":
                self.view_transactions()
            elif choice == "5":
                self.save_users()
                self.save_transactions()
                print("Exiting... Goodbye!")
                break
            else:
                print("Invalid option. Please try again.")


if __name__ == "__main__":
    system = MobilePaymentSystem()
    system.run()
