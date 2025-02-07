import openpyxl
from datetime import datetime

# --- Classes --- Create------

class Wallet:
    def __init__(self, balance=0.0):
        self.balance = balance

    def deposit(self, amount):
        self.balance += amount

    def withdraw(self, amount):
        if self.balance >= amount:
            self.balance -= amount
            return True
        return False

    def check_balance(self):
        return self.balance


class User:
    def __init__(self, user_id, name, phone_number, wallet=Wallet()):  # Wallet object creation
        self.user_id = user_id
        self.name = name
        self.phone_number = phone_number
        self.wallet = wallet

    def display_details(self):
        return f"User ID: {self.user_id}, Name: {self.name}, Phone: {self.phone_number}"

    def receive_money(self, amount):
        self.wallet.deposit(amount)  # Money received and added to wallet


class Transaction:
    def __init__(self, transaction_id, sender, receiver, amount, date=None):
        self.transaction_id = transaction_id
        self.sender = sender
        self.receiver = receiver
        self.amount = amount
        self.date = f"{datetime.now():%Y-%m-%d %H:%M:%S}"

    def to_dict(self):
        return {
            "Transaction ID": self.transaction_id,
            "Sender": self.sender.user_id,
            "Receiver": self.receiver.user_id,
            "Amount": self.amount,
            "Date": self.date,
        }


# --- Excel Create & Save Functions ---

def load_last_user_id(filename="users.xlsx"):
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        last_row = sheet.max_row
        if last_row > 1:  # Check if there are any users in the file
            last_user_id = sheet.cell(row=last_row, column=1).value
            if last_user_id.startswith("U"):
                return int(last_user_id[1:])
    except FileNotFoundError:
        pass
    return 0  # Default if no file or users exist


def save_users(users, filename="users.xlsx"):
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["User ID", "Name", "Phone Number", "Balance"])
        wb.save(filename)
        wb = openpyxl.load_workbook(filename)

    sheet = wb.active
    existing_ids = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True)}

    for user in users.values():
        if user.user_id not in existing_ids:
            sheet.append([user.user_id, user.name, user.phone_number, user.wallet.check_balance()])
            existing_ids.add(user.user_id)

    wb.save(filename)


def save_transactions(transactions, filename="transactions.xlsx"):
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Transaction ID", "Sender ID", "Receiver ID", "Amount", "Date"])
        wb.save(filename)
        wb = openpyxl.load_workbook(filename)

    sheet = wb.active
    existing_ids = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True)}

    for t in transactions:
        if t.transaction_id not in existing_ids:
            sheet.append([t.transaction_id, t.sender.user_id, t.receiver.user_id, t.amount, t.date])
            existing_ids.add(t.transaction_id)

    wb.save(filename)


# --- Main Program ---

def main():
    users = {}
    transactions = []
    last_user_id = load_last_user_id()

    print("\nWelcome to the Mobile Payment System!")

    while True:
        print("1. Register New User")
        print("2. Check Balance")
        print("3. Send Money")
        print("4. Receive Money")
        print("5. View Transaction History")
        print("6. Exit")
        choice = input("Choose an option: ")

        if choice == "1":
            name = input("Enter Name: ")
            phone = input("Enter Phone Number: ")
            initial_balance = float(input("Enter Initial Balance: "))
            last_user_id += 1
            user_id = f"U{last_user_id:03d}"  # Auto-increment User ID
            users[user_id] = User(user_id, name, phone, Wallet(initial_balance))
            save_users(users)
            print(f"User registered successfully with User ID: {user_id}!")

        elif choice == "2":
            user_id = input("Enter User ID: ")
            if user_id in users:
                balance = users[user_id].wallet.check_balance()
                print(f"Current Balance: ${balance:.2f}")
            else:
                print("User not found.")

        elif choice == "3":
            sender_id = input("Enter Sender User ID: ")
            receiver_id = input("Enter Receiver User ID: ")
            amount = float(input("Enter Amount: "))
            if sender_id in users and receiver_id in users:
                sender = users[sender_id]
                receiver = users[receiver_id]
                if sender.wallet.withdraw(amount):
                    receiver.receive_money(amount)
                    transaction_id = f"T{len(transactions)+1:03d}"
                    transaction = Transaction(transaction_id, sender, receiver, amount)
                    transactions.append(transaction)
                    save_users(users)  # Save updated balances
                    print(f"Transaction successful! {amount:.2f} sent to {receiver.name}.")
                else:
                    print("Insufficient balance. Transaction failed.")
            else:
                print("Sender or Receiver not found.")

        elif choice == "4":
            user_id = input("Enter User ID: ")
            receiver_id = input("Enter Receiver User ID: ")
            amount = float(input("Enter Amount: "))
            if user_id in users and receiver_id in users:
                users[user_id].receive_money(amount)
                save_users(users)  # Save updated balance
                print(f"Transaction successful! ${amount:.2f} received.")
            else:
                print("User not found.")

        elif choice == "5":
            print("Transaction History:")
            for t in transactions:
                print(f"{t.sender.user_id} sent ${t.amount:.2f} to {t.receiver.user_id} on {t.date}")

        elif choice == "6":
            save_users(users)
            save_transactions(transactions)
            print("Exiting... Goodbye!")
            break

        else:
            print("Invalid option. Please try again.")


if __name__ == "__main__":
    main()
