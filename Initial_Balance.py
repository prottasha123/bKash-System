import openpyxl
from datetime import datetime

# --- Classes --- 

class Wallet:
    def __init__(self, balance=0.0):
        self.balance = balance

    def deposit(self, amount):
        self.balance += amount

    def withdraw(self, amount):
        if self.balance >= amount:
            self.balance -= amount
            return True
        return False

    def check_balance(self):
        return self.balance


class User:
    def __init__(self, user_id, name, phone_number, wallet=Wallet()):  # Object creation Wallet class
        self.user_id = user_id
        self.name = name
        self.phone_number = phone_number
        self.wallet = wallet

    def display_details(self):
        return f"User ID: {self.user_id}, Name: {self.name}, Phone: {self.phone_number}"


class Transaction:
    def __init__(self, transaction_id, sender, receiver, amount, date=None):  # Date initial value None
        self.transaction_id = transaction_id
        self.sender = sender
        self.receiver = receiver
        self.amount = amount
        self.date = f"{datetime.now():%Y-%m-%d %H:%M:%S}"

    def to_dict(self):
        return {
            "Transaction ID": self.transaction_id,
            "Sender": self.sender.user_id,
            "Receiver": self.receiver.user_id,
            "Amount": self.amount,
            "Date": self.date,
        }


# --- Excel Create & Save Functions --- 

def save_users(users, filename="users.xlsx"):
    wb = openpyxl.Workbook()  # Create a new workbook
    sheet = wb.active  # Get the active worksheet
    sheet.append(["User ID", "Name", "Phone Number", "Balance"])  # Add headers to the sheet

    for user in users.values():  # Iterate over the values in the dictionary
        sheet.append([user.user_id, user.name, user.phone_number, user.wallet.check_balance()])
    wb.save(filename)


def save_transactions(transactions, filename="transactions.xlsx"):
    # Load the existing file or create a new one
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Transaction ID", "Sender ID", "Receiver ID", "Amount", "Date"])
        wb.save(filename)
        wb = openpyxl.load_workbook(filename)

    sheet = wb.active

    # Get the existing transaction IDs to avoid duplicates
    existing_ids = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True)}  # Set comprehension

    # Get the last transaction ID and increment it for new transactions
    last_transaction_id = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        transaction_id = row[0]
        if transaction_id.startswith("T"):
            try:
                last_transaction_id = max(last_transaction_id, int(transaction_id[1:]))
            except ValueError:
                pass

    # Append new transactions with auto-incremented transaction IDs
    for t in transactions:
        last_transaction_id += 1
        new_transaction_id = f"T{last_transaction_id:03d}"

        if new_transaction_id not in existing_ids:
            sheet.append([new_transaction_id, t.sender.user_id, t.receiver.user_id, t.amount, t.date])
            existing_ids.add(new_transaction_id)

    # Save the updated file
    wb.save(filename)


# --- Main Program --- 

def main():
    users = {}  # Initialize an empty dictionary to store users
    transactions = []  # Initialize an empty list to store transactions

    print("\nWelcome to the Mobile Payment System!")

    while True:
        print("1. Register New User")
        print("2. Check Balance")
        print("3. Send Money")
        print("4. Receive Money")  # New option added
        print("5. View Transaction History")
        print("6. Exit")
        choice = input("Choose an option: ")

        if choice == "1":
            user_id = input("Enter User ID: ")
            name = input("Enter Name: ")
            phone = input("Enter Phone Number: ")
            balance = float(input("Enter Initial Balance: "))
            if user_id in users:
                print("User ID already exists.")
            else:
                users[user_id] = User(user_id, name, phone, Wallet(balance))
                save_users(users)
                print("User registered successfully!")

        elif choice == "2":
            user_id = input("Enter User ID: ")
            if user_id in users:
                balance = users[user_id].wallet.check_balance()
                print(f"Current Balance: ${balance:.2f}")
            else:
                print("User not found.")

        elif choice == "3":
            sender_id = input("Enter Sender User ID: ")
            receiver_id = input("Enter Receiver User ID: ")
            amount = float(input("Enter Amount: "))
            if sender_id == receiver_id:
                print("Error: Sender and Receiver cannot be the same.")
            elif sender_id in users and receiver_id in users:
                sender = users[sender_id]
                receiver = users[receiver_id]

                if sender.wallet.check_balance() < amount:
                    print("Error: Insufficient balance.")
                elif sender.wallet.withdraw(amount):
                    receiver.wallet.deposit(amount)
                    transaction_id = f"T{len(transactions)+1:034d}"
                    transaction = Transaction(transaction_id, sender, receiver, amount)
                    transactions.append(transaction)
                    save_transactions(transactions)
                    print(f"Transaction successful! {amount:.2f} sent to {receiver.name}.")
                else:
                    print("Error: Transaction failed.")
            else:
                print("Error: Sender or Receiver not found.")

        elif choice == "4":  # New case for receiving money
            receiver_id = input("Enter Receiver User ID: ")
            sender_id = input("Enter Sender User ID: ")
            amount = float(input("Enter Amount to Receive: "))

            if sender_id == receiver_id:
                print("Error: Sender and Receiver cannot be the same.")
            elif sender_id in users and receiver_id in users:
                sender = users[sender_id]
                receiver = users[receiver_id]

                if sender.wallet.withdraw(amount):  # Sender must have sufficient balance
                    receiver.wallet.deposit(amount)
                    transaction_id = f"T{len(transactions)+1:03d}"
                    transaction = Transaction(transaction_id, sender, receiver, amount)
                    transactions.append(transaction)
                    save_transactions(transactions)
                    print(f"{amount:.2f} successfully received from {sender.name}.")
                else:
                    print("Error: Sender has insufficient funds.")
            else:
                print("Error: Sender or Receiver not found.")

        elif choice == "5":
            if transactions:  # Check if the transactions list is not empty
                print("Transaction History:")
                for t in transactions:
                    print(f"{t.sender.user_id} sent ${t.amount:.2f} to {t.receiver.user_id} on {t.date}")
            else:
                print("No transactions found.")


        elif choice == "6":
            print("Exiting... Goodbye!")
            break

        else:
            print("Invalid option. Please try again.")


if __name__ == "__main__":
    main()
