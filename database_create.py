
'''
from openpyxl import Workbook, load_workbook


def open_xlDatabase():
    db = Workbook()
    db_sheet = db.active
    db_sheet.title = "User Infromation"
    db.save("users.xlsx")


open_xlDatabase()


def insert_login_data(userid, username, phone_number,balance):
  db=load_workbook("users.xlsx")
  db_sheet=db.active
  db_sheet.title="Login infromation"
  db_sheet.append([userid, username, phone_number,balance])
  db.save("users.xlsx")

insert_login_data("User ID","Name","Phone Number","Balance")

insert_login_data("U001", "Mahmuda Islam", "01712345678",1000)
insert_login_data("U002", "John Doe", "01898765432",2000)
insert_login_data("U003", "Jane Smith", "01987654321",3000)
insert_login_data("U004", "Alice Brown", "01623456789",4000)
insert_login_data("U005", "Bob White", "01512349876",5000)
insert_login_data("U006", "Charlie Green", "01787654321",6000)
insert_login_data("U007", "Emily Blue", "01876543210",7000)
insert_login_data("U008", "David Red", "01965432109",8000)
insert_login_data("U009", "Grace Yellow", "01654321098",9000)
insert_login_data("U010", "Olivia Black", "01543210987",1000)

'''

'''

from openpyxl import Workbook, load_workbook


def open_xlDatabase():
    db = Workbook()
    db_sheet = db.active
    db_sheet.title = "Transaction Infromation"
    db.save("transactions.xlsx")


open_xlDatabase()



def insert_login_data(Transaction_ID, Sender_ID, Receiver_ID,Amount,Date):
  db=load_workbook("transactions.xlsx")
  db_sheet=db.active
  db_sheet.title="Login infromation"
  db_sheet.append([Transaction_ID, Sender_ID, Receiver_ID,Amount,Date])
  db.save("transactions.xlsx")

insert_login_data("Transaction ID","Sender ID","Receiver ID","Amount","Date")


'''

