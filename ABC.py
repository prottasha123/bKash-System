from openpyxl import Workbook, load_workbook
from datetime import datetime

# Vehicle class
class Vehicle:
    def __init__(self, vehicle_id, vehicle_type, license_plate):
        self.vehicle_id = vehicle_id
        self.vehicle_type = vehicle_type
        self.license_plate = license_plate

    def display_vehicle_details(self):
        return f"ID: {self.vehicle_id}, Type: {self.vehicle_type}, Plate: {self.license_plate}"


# TollBooth class
class TollBooth:
    def __init__(self, booth_id, location):
        self.booth_id = booth_id
        self.location = location
        self.toll_rates = {}

    def set_toll_rates(self, rates):
        self.toll_rates = rates
        

    def calculate_toll(self, vehicle_type):
        return self.toll_rates.get(vehicle_type, 0)


# TollTransaction class
class TollTransaction:
    def __init__(self, transaction_id, vehicle, toll_booth, amount):
        self.transaction_id = transaction_id
        self.vehicle = vehicle
        self.toll_booth = toll_booth
        self.amount = amount
        self.timestamp = datetime.now()

    def display_transaction_details(self):
        return f"ID: {self.transaction_id}, Vehicle: {self.vehicle.vehicle_id}, Booth: {self.toll_booth.booth_id}, " \
               f"Amount: ${self.amount:.2f}, Date: {self.timestamp.strftime('%Y-%m-%d %H:%M:%S')}"


# TollManagementSystem class
class TollManagementSystem:
    def __init__(self):
        self.vehicles = []
        self.toll_booths = []
        self.transactions = []
        self.vehicle_file = "vehicles.xlsx"
        self.booth_file = "toll_booths.xlsx"
        self.transaction_file = "transactions.xlsx"
        self.load_data()

    def load_data(self):
        # Load vehicles
        try:
            wb = load_workbook(self.vehicle_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                vehicle = Vehicle(row[0], row[1], row[2])
                self.vehicles.append(vehicle)
            wb.close()
        except FileNotFoundError:
            print("No existing vehicle data found. Starting fresh.")

        # Load toll booths
        try:
            wb = load_workbook(self.booth_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                booth = TollBooth(row[0], row[1])
                booth.set_toll_rates(eval(row[2]))
                print(eval(row[2]))
                self.toll_booths.append(booth)
            wb.close()
            
            print(eval(row[2]))
            print(row)
            
        except FileNotFoundError:
            print("No existing toll booth data found. Starting fresh.")

        # Load transactions
        try:
            wb = load_workbook(self.transaction_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                vehicle = next((v for v in self.vehicles if v.vehicle_id == row[1]), None)
                toll_booth = next((b for b in self.toll_booths if b.booth_id == row[2]), None)
                transaction = TollTransaction(row[0], vehicle, toll_booth, row[3])
                transaction.timestamp = datetime.strptime(row[4], '%Y-%m-%d %H:%M:%S')
                self.transactions.append(transaction)
            wb.close()
        except FileNotFoundError:
            print("No existing transaction data found. Starting fresh.")

    def save_data(self):
        # Save vehicles
        wb = Workbook()
        ws = wb.active
        ws.append(["Vehicle ID", "Vehicle Type", "License Plate"])
        for vehicle in self.vehicles:
            ws.append([vehicle.vehicle_id, vehicle.vehicle_type, vehicle.license_plate])
        wb.save(self.vehicle_file)

        # Save toll booths
        wb = Workbook()
        ws = wb.active
        ws.append(["Booth ID", "Location", "Toll Rates"])
        for booth in self.toll_booths:
            ws.append([booth.booth_id, booth.location, str(booth.toll_rates)])
        wb.save(self.booth_file)

        # Save transactions
        wb = Workbook()
        ws = wb.active
        ws.append(["Transaction ID", "Vehicle ID", "Booth ID", "Amount", "Timestamp"])
        for transaction in self.transactions:
            ws.append([transaction.transaction_id, transaction.vehicle.vehicle_id, transaction.toll_booth.booth_id,
                       transaction.amount, transaction.timestamp.strftime('%Y-%m-%d %H:%M:%S')])
        wb.save(self.transaction_file)

    def add_vehicle(self, vehicle_id, vehicle_type, license_plate):
        if any(vehicle.vehicle_id == vehicle_id for vehicle in self.vehicles):
            print("Vehicle with this ID already exists. Cannot add duplicate.")
            return

        vehicle = Vehicle(vehicle_id, vehicle_type, license_plate)
        self.vehicles.append(vehicle)
        print("Vehicle added successfully!")

    def add_toll_booth(self, booth_id, location, toll_rates):
        if any(booth.booth_id == booth_id and booth.location == location for booth in self.toll_booths):
            print("Toll booth with this ID already exists in this location. Cannot add duplicate.")
            return

        booth = TollBooth(booth_id, location)
        booth.set_toll_rates(toll_rates)
        self.toll_booths.append(booth)
        print("Toll booth added successfully!")

    def record_transaction(self, transaction_id, vehicle_id, booth_id):
        vehicle = next((v for v in self.vehicles if v.vehicle_id == vehicle_id), None)
        toll_booth = next((b for b in self.toll_booths if b.booth_id == booth_id), None)
        if vehicle and toll_booth:
            amount = toll_booth.calculate_toll(vehicle.vehicle_type)
            transaction = TollTransaction(transaction_id, vehicle, toll_booth, amount)
            self.transactions.append(transaction)
            return f"Transaction recorded successfully! Amount: ${amount:.2f}"
        return "Invalid Vehicle ID or Booth ID."

    def view_transaction_history(self):
        if not self.transactions:
            return "No transactions found."
        return "\n".join(t.display_transaction_details() for t in self.transactions)


# Sample Menu Interface
def main():
    tms = TollManagementSystem()
    print("\nWelcome to the Toll Management System!")

    while True:
        print("1. Add Vehicle")
        print("2. Add Toll Booth")
        print("3. Record Toll Transaction")
        print("4. View Transaction History")
        print("5. Exit")
        choice = input("Choose an option: ")

        if choice == '1':
            vehicle_id = input("Enter Vehicle ID: ")
            vehicle_type = input("Enter Vehicle Type (Car/Truck/Motorcycle): ")
            license_plate = input("Enter License Plate: ")
            tms.add_vehicle(vehicle_id, vehicle_type, license_plate)
            tms.save_data()

        elif choice == '2':
          booth_id = input("Enter Toll Booth ID: ")
          location = input("Enter Location: ")
          rates = input("Enter Toll Rates (e.g., {'Car': 5, 'Truck': 10, 'Motorcycle': 3}): ")
          rates_dict = eval(rates)
          print(f"Toll Rates Input: {rates_dict}")  # Print the dictionary here
          tms.add_toll_booth(booth_id, location, rates_dict)
          tms.save_data()


        elif choice == '3':
            transaction_id = input("Enter Transaction ID: ")
            vehicle_id = input("Enter Vehicle ID: ")
            booth_id = input("Enter Toll Booth ID: ")
            print(tms.record_transaction(transaction_id, vehicle_id, booth_id))
            tms.save_data()

        elif choice == '4':
            print("\nTransaction History:")
            print(tms.view_transaction_history())

        elif choice == '5':
            tms.save_data()
            print("Exiting... Goodbye!")
            break

        else:
            print("Invalid option. Please try again.")


if __name__ == "__main__":
    main()
       
